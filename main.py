from openpyxl import load_workbook
import json
from neo4j import GraphDatabase
from fastapi import FastAPI
import re

file_way = 'tvel.xlsx'
wb = load_workbook(file_way)
sheet_mk = wb['Мультикубы - Кубы']
if 'recoding' in wb.sheetnames:
    wb.remove(wb['recoding'])
wb.create_sheet('recoding')
sheet_encoding = wb['recoding']
row_counter = 0
_dct = {}
dct_with_var = {}
result_dct = {}
archive = set()
row_encoding = 1


def make_var(string):
    string = re.sub(r'[^a-zA-Zа-яА-ЯёЁ0-9]', r'_', string.strip('\'')).strip('_').strip('\'')
    if string != '':
        if string[0].isdigit():
            string = '_' + string
    string = re.sub(r'_+', '_', string)
    return string


def create_dict(_list, parent):
    global row_encoding
    for item in _list:
        temp_key = make_var(item)
        temp_value = make_var(parent)
        if '\'.' not in item:
            mk = find_mk()
            mk = make_var(mk)
            temp_key = re.sub(r'_+', '_', mk + '_' + temp_key)
        temp_value = re.sub(r'_+', '_', temp_value)
        if temp_key in _dct.keys():
            temp_key = re.sub(r'_+', '_', temp_key)
            _dct[temp_key].append(temp_value)

        else:
            temp_key = re.sub(r'_+', '_', temp_key)
            _dct[temp_key] = [temp_value]
        sheet_encoding.cell(row=row_encoding, column=1).value = item
        sheet_encoding.cell(row=row_encoding, column=2).value = temp_key
        row_encoding += 1
        sheet_encoding.cell(row=row_encoding, column=1).value = parent
        sheet_encoding.cell(row=row_encoding, column=2).value = temp_value
        row_encoding += 1


def find_mk():
    for i in range(row_counter, 1, -1):
        if sheet_mk.cell(row=i, column=2).value is None:
            return sheet_mk.cell(row=i, column=1).value


class App:

    def __init__(self, uri, user, password):
        self.driver = GraphDatabase.driver(uri, auth=(user, password))

    def close(self):
        self.driver.close()

    def create_request(self, lst):
        with self.driver.session(database="neo4j") as session:
            result = session.execute_write(
                self._create_and_return_relation, lst)

    @staticmethod
    def _create_and_return_relation(tx, dct):
        query = ''
        for key in dct.keys():
            if key not in archive:
                if query == '':
                    query = f'({key}:Cube {{ name: "{key}"}})'
                    archive.add(key)
                else:
                    query = ', '.join([f'({key}:Cube {{ name: "{key}"}})', query.strip(', ')])
                    archive.add(key)
            for child in dct[key]:
                if child not in archive:
                    query = ', '.join([f'({child}:Cube {{ name: "{child}"}})', query.strip(', ')])
                    archive.add(child)
                if f'({child})-[:RELATION]->({key})' not in query:
                    make_relation = f'({child})-[:RELATION]->({key})'
                else:
                    make_relation = ''
                if query != '':
                    query = ', '.join([query.strip(', '), make_relation.strip(', ')])
                else:
                    query = make_relation
        with open("file.txt", "w") as output:
            output.write(str('create ' + re.sub(r'(,\s*,)+', ',', query).replace(', ,', ',').replace(',  ,', ',')
                             .replace(',   ,', ',').replace(',    ,', ',')))
        result = tx.run(
            ('create ' + re.sub(r'(,\s*,)+', ',', query)).replace(', ,', ',').replace(',  ,', ',')
            .replace(',   ,', ',').replace('    ', ',').replace(', ,', ',').strip())
        return [result]

    app = FastAPI()

    @app.get("/match")
    def return_name_json(self, name):
        session = self.driver.session()
        query = f'''MATCH (n:Cube {{name:"{name}"}}) RETURN {{elementId:
        elementId(n), labels: LABELs(n), properties: PROPERTies(n)}} as query'''
        result = json.dumps(session.run(query).data())
        decoding_result = json.loads(result)
        return decoding_result


if __name__ == '__main__':
    for line in sheet_mk.values:
        row_counter += 1
        lst = list(line)
        if lst[0] is not None and lst[2] is not None:
            result_lst = lst[2].split(', ')
            for r in range(row_counter, 1, -1):
                if sheet_mk.cell(row=r, column=2).value is None:
                    parent_mk_cube = sheet_mk.cell(row=r, column=1).value + '.' + sheet_mk.cell(row=row_counter,
                                                                                                column=1).value
                    create_dict(result_lst, parent_mk_cube)
                    break

wb.save(file_way)
wb.close()
print('Start')
uri = "neo4j+s://7e0c70fc.databases.neo4j.io"
user = "neo4j"
password = "oaYU6dMmKz7G2GgUJGRjDbh32mSfVxbAQiZZo-cODrY"
app = App(uri, user, password)
app.create_request(_dct)
app.close()
print('End')
